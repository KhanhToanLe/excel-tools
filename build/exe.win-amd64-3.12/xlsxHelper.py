from openpyxl import *
from os import listdir
from os.path import isfile, join, exists
from tkinter import *
import re
import InputData as data
import xlwings as xw
import time
import math
from re import match
import copy
import traceback
import exception_logger as exlogger
import helper

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

merge_sheet_prefix = helper.PREFIX_MERGE

def get_files(directory):
	directory = directory.replace("\\","/")
	files = [f for f in listdir(directory) if isfile(join(directory, f)) and f.endswith(".xlsx")]
	files = [directory + f"/{f}" for f in files]
	return files

def remove_message_box(top_level, after_action=None):
	top_level.destroy()
	top_level.update()
	if after_action == None: return
	after_action()

def add_message_box(message,parent_frame,width=150,height=120, unmount_action=None):
	top= Toplevel(parent_frame,height=height,width=width)
	x = parent_frame.winfo_x()
	y = parent_frame.winfo_y()
	top.geometry("%dx%d+%d+%d" % (width, height, x + 200, y + 200))
	top.title("Info")
	Label(top, text= message, font=('Arial 11 normal'),pady=20).pack()
	Button(top, text= "Ok", font=('Arial 11 normal'),command=lambda:remove_message_box(top,unmount_action)).pack()

def validate_directory(directory):
	if directory == "" or not exists(directory):
		return False
	return True

def validate_directory_and_select_sheet(directory, select_sheet, window):
	is_valid_directory = validate_directory(directory)
	if not is_valid_directory :
		add_message_box("Directory is invalid!",window)
		return False
	try: 
		select_sheet = int(select_sheet)
		select_sheet = 0 if select_sheet - 1 <= 0 else select_sheet - 1
	except:
		if select_sheet != data.INFINITE_SHEET_CHECKER:
			add_message_box("Invalid sheet number",window)
			exlogger.write(traceback.format_exc())
			return False
	return True
	
def change_data(command_list, directory,window, select_sheet):
	is_valid = validate_directory_and_select_sheet(directory,select_sheet,window)
	if not is_valid: 
		return
	files = get_files(directory)
	for file in files:
		try:
			wb = load_workbook(filename = file)
			worksheets = []
			if select_sheet == data.INFINITE_SHEET_CHECKER:
				worksheets = wb.worksheets
			else:
				select_sheet = int(select_sheet)
				select_sheet = 0 if select_sheet - 1 <= 0 else select_sheet - 1
				worksheets = [wb.worksheets[select_sheet]]
			for worksheet in worksheets:
				for command in command_list: 
					update_cell = command.cell
					is_match = re.match(r"^[a-zA-Z]+\d+$",update_cell)
					if not is_match:
						if update_cell == "":
							add_message_box("Empty cell value",window)
							return
						add_message_box(f"'{update_cell}' isn't a cell",window)
						return
					update_value = command.value
					worksheet[update_cell] = update_value
			wb.save(file)
		except Exception as ex:
			# exception: index out of range
			if type(ex) ==  IndexError:
				file_name_log = file.split("/")[-1]
				add_message_box(f"Error {file_name_log} \n out of range",window,height=140,width=150)
				return
			if type(ex) == PermissionError:
				add_message_box("Turn off all of\n .xlsx files first",window,height=140,width=150)
				return
			# exception: all
			add_message_box("Error detected!",window)
			exlogger.write(traceback.format_exc())
			return
	add_message_box("Execute completed!",window)

def get_column_by_index(index):
	global alphabet
	divide_val = int(index/26)-1
	offset_val = int(index%26)-1
	if divide_val == -1:
		return alphabet[offset_val]
	return f"{alphabet[divide_val]}{alphabet[offset_val]}"

def control_home(directory, select_sheet, window,process_bar_increase_val):
	is_valid = validate_directory_and_select_sheet(directory,select_sheet, window)
	if not is_valid: 
		return
	files = get_files(directory)
	try:
		app = xw.App(visible=True)
		steps = 100/len(files)
		
		for file in files:
			book = xw.Book(file)
			active_window = book.app.api.ActiveWindow
			for sheet in book.sheets:
				sheet.activate()
				is_freeze_panes = active_window.FreezePanes

				if is_freeze_panes:
					freeze_row =  int(active_window.SplitRow) + 1
					freeze_column = int(active_window.SplitColumn) + 1
					freeze_column = get_column_by_index(freeze_column)
					select_coord = f"{freeze_column}{freeze_row}"
					sheet.range(select_coord).select()
				else:
					sheet.range("A1").select()
			book.sheets[0].activate()
			book.save(file)
			book.close()
			process_bar_increase_val(steps)
		add_message_box("Execute completed",window,unmount_action=lambda:process_bar_increase_val(0))
		app.quit()

	except Exception as ex:
		add_message_box("Error detected!",window)
		exlogger.write(traceback.format_exc())

def get_max_number(strings):
	max_number = None
	global merge_sheet_prefix
	pattern_regex = rf"\b{merge_sheet_prefix}(\d+)\b"

	for string in strings:
		match = re.search(pattern_regex, string)
		if match:
			number = int(match.group(1))
			if max_number is None or number > max_number:
				max_number = number

	if max_number == None: 
		return 0
	return max_number

def check_if_merge_index_of_sheet(choose_sheet):
	try:
		select_sheet = int(choose_sheet)
		select_sheet = 0 if select_sheet - 1 <= 0 else select_sheet - 1
		return select_sheet
	except Exception as ex:
		return "Not Set"

def delete_old_merge_sheet(destination_work_book_directory,destination_work_book:Workbook,match_list):
	for match_item in match_list:
		destination_work_book.remove(destination_work_book[match_item])
	destination_work_book.save(destination_work_book_directory);

def merge_sheet(directory, choose_sheet, window, destination_file, is_remove_old_merge_sheet):
	try:
		is_valid = validate_directory_and_select_sheet(directory,choose_sheet, window)
		if not is_valid: 
			return
		files = get_files(directory)
		destination_file = destination_file.replace("\\","/")
		if destination_file == "" or not isfile(destination_file):
			add_message_box("Invalid destination file path",window,width=200)
			return

		destination_work_book = load_workbook(filename=destination_file)
		destination_work_book_sheet_book_list = destination_work_book.sheetnames
		global merge_sheet_prefix
		match_list = list(filter(lambda v: match(rf'\b{merge_sheet_prefix}\d+\b', v), destination_work_book_sheet_book_list))
		if is_remove_old_merge_sheet:
			delete_old_merge_sheet(destination_file,destination_work_book, match_list)
			match_list= []
		maxValue = get_max_number(match_list)
		sheet_merge_name_index = maxValue
		merge_index_sheet = check_if_merge_index_of_sheet(choose_sheet)

		for files  in files:
			if files == destination_file: continue
			# sheet you want to copy
			source = load_workbook(files)
			sheet_list = source.sheetnames
			sheet_list_loop = None
			if merge_index_sheet == "Not Set":
				sheet_list_loop = sheet_list
			else:
				sheet_list_loop = [sheet_list[merge_index_sheet]]

			for sheet in sheet_list_loop:
				first_sheet = source[sheet]
				sheet_merge_name_index = sheet_merge_name_index + 1
				global merge_sheet_prefix
				first_sheet.title = merge_sheet_prefix + str(sheet_merge_name_index)
				first_sheet._parent = destination_work_book
				destination_work_book._add_sheet(first_sheet)
			source.close()
		destination_work_book.save(destination_file)
		destination_work_book.close()
		
		app = xw.App(visible=True)
		destination_book = app.books.open(destination_file)
		sheet_length = len(destination_book.sheets)

		for sheet in reversed(range(0,int(math.ceil(sheet_length)))):
			destination_book.sheets[sheet].activate()
		destination_book.save()
		app.quit()
		add_message_box("Execute completed", window)
		
	except Exception as ex:
		if isinstance(ex,PermissionError):
			add_message_box("Turn off all .xlsx files first ",window,height=200)
			return
		add_message_box("Error Detect!",window)
		print(ex)
		return