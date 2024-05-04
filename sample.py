from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
wb = load_workbook(filename='test/freeze_both.xlsx')
ws = wb['Test case']


def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col) if len(rng) != 0 else cell


cell1 = ws["A2"]
# cell1 = sheet_ranges.cell(row=15,column=14)
# ws.cell(ws["A4"]).value = "Hello World"
test = getMergedCellVal(ws, cell1)
print(test)
# print(cell1.)
# print(type(cell1))
# if isinstance(cell1, MergedCell):
#     print(cell1.coordinate)
#     print("Oh no, the cell is merged!")
# else:
#     print("This cell is not merged.")
